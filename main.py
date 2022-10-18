from openpyxl import Workbook, load_workbook

wb_path = "automatizador_planilhas/planilhas/tabela_cliente.xlsx"
wb = load_workbook(wb_path)

ws = wb.active

max_linha = ws.max_row
max_coluna = ws.max_column

for i in range(2, max_linha + 1):
  for j in range(2, max_coluna + 1):
    print(ws.cell(row=i, column=j).value, end=" - ")


